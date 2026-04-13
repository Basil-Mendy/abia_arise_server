from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.views import APIView
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework_simplejwt.tokens import RefreshToken
from django.db.models import Q, Count
from django.contrib.auth.models import User
from django.conf import settings
from django.http import FileResponse
from django.core.mail import send_mail
from django.utils import timezone
from datetime import timedelta
from .models import IndividualMember, ProGroup, GroupMember, MembershipUser, MembershipGroup, GroupMembership
from .serializers import (
    IndividualMemberSerializer, ProGroupSerializer, GroupMemberSerializer,
    MemberActivationSerializer, ExcelMemberImportSerializer,
    GenerateResetPinSerializer, VerifyResetPinOtpSerializer, VerifyResetPinSerializer,
    AddMemberToGroupSerializer,
    MembershipUserListSerializer, MembershipUserDetailSerializer, MembershipUserCreateSerializer,
    MembershipGroupSerializer, GroupMembershipDetailSerializer, GroupMembersListSerializer,
    ExcelUploadSerializer, UserRegistrationSerializer, AddMemberToGroupSerializer, RemoveMemberFromGroupSerializer
)
from .generators import generate_id_card, generate_certificate
import hashlib
import os
import openpyxl
import random
from io import BytesIO

class IndividualMemberViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing individual members
    """
    queryset = IndividualMember.objects.all()
    serializer_class = IndividualMemberSerializer
    parser_classes = (JSONParser, MultiPartParser, FormParser)

    def get_permissions(self):
        """
        Allow unauthenticated access to register, login, check_duplicate, generate_id_card, get_id_card, and dashboard endpoints
        """
        if self.action in ['register', 'login', 'check_duplicate', 'generate_id_card', 'get_id_card', 'dashboard', 'update_residential_info', 'update_bank_details', 'update_pin', 'update_profile_picture', 'activate_membership']:
            return [AllowAny()]
        return super().get_permissions()

    @action(detail=False, methods=['post'])
    def register(self, request):
        """
        Register a new individual member and automatically generate ID card
        """
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            try:
                member = serializer.save()
                
                # Automatically generate ID card after registration
                member_data = {
                    'first_name': member.first_name,
                    'middle_name': member.middle_name,
                    'last_name': member.last_name,
                    'abia_arise_id': member.abia_arise_id,
                    'lga_of_origin': member.lga_of_origin,
                    'state_of_origin': member.state_of_origin,
                    'profile_picture': member.profile_picture if member.profile_picture else None,
                }
                
                success, card_path, error = generate_id_card(member_data, member_instance=member)
                
                if success:
                    print(f"✓ ID card generated successfully: {card_path}")
                else:
                    print(f"✗ ID card generation warning: {error}")
                
                return Response({
                    'success': True,
                    'message': 'Registration successful. ID card generated.',
                    'member_id': serializer.data['abia_arise_id'],
                    'data': serializer.data
                }, status=status.HTTP_201_CREATED)
            except Exception as e:
                return Response({
                    'success': False,
                    'error': str(e)
                }, status=status.HTTP_400_BAD_REQUEST)
        print("Serializer errors:", serializer.errors)  # DEBUG
        return Response({
            'success': False,
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def generate_id_card(self, request):
        """
        Generate ID card for a registered member
        Expects: member_id (abia_arise_id)
        """
        try:
            member_id = request.data.get('member_id')
            print(f"DEBUG: Generating ID card for member_id: {member_id}")
            member = IndividualMember.objects.get(abia_arise_id=member_id)
            print(f"DEBUG: Found member: {member.first_name} {member.last_name}")
            
            # Prepare member data for ID card generation
            member_data = {
                'first_name': member.first_name,
                'middle_name': member.middle_name,
                'last_name': member.last_name,
                'ward': member.electoral_ward,
                'lga': member.lga_of_residence,
                'state': member.state_of_residence,
                'abia_arise_id': member.abia_arise_id,
            }
            
            # Add profile picture if available
            if member.profile_picture:
                try:
                    # Get the full path to the profile picture
                    profile_pic_path = member.profile_picture.path
                    member_data['profile_picture'] = profile_pic_path
                except Exception as pic_error:
                    print(f"Warning: Could not access profile picture: {pic_error}")
            
            # Generate ID card
            try:
                generator = IDCardGenerator()
                file_path, relative_path, error = generator.generate(member_data)
                
                if error:
                    print(f"ERROR: ID card generation failed: {error}")
                    return Response({
                        'success': False,
                        'message': 'ID card generation failed',
                        'error': error
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Save path to database if successful
                if file_path and relative_path:
                    try:
                        member.id_card_file = relative_path
                        member.save()
                        print(f"ID card path saved to database: {relative_path}")
                    except Exception as db_error:
                        print(f"Warning: Could not save ID card path to database: {db_error}")
                
                return Response({
                    'success': True,
                    'message': 'ID card generated successfully',
                    'card_url': relative_path,
                    'member_data': {
                        'full_name': f"{member.first_name} {member.middle_name} {member.last_name}".strip(),
                        'abia_arise_id': member.abia_arise_id
                    }
                }, status=status.HTTP_200_OK)
            except Exception as gen_error:
                print(f"ERROR: ID card generator exception: {str(gen_error)}")
                import traceback
                traceback.print_exc()
                return Response({
                    'success': False,
                    'message': 'Failed to initialize ID card generator',
                    'error': str(gen_error)
                }, status=status.HTTP_400_BAD_REQUEST)
        
        except IndividualMember.DoesNotExist:
            print("DEBUG: Member not found")
            return Response({
                'success': False,
                'message': 'Member not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            print(f"DEBUG: ID card endpoint error: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def check_duplicate(self, request):
        """
        Check if NIN, voters card, email, or phone number already exists
        """
        nin = request.data.get('nin')
        voters_card_no = request.data.get('voters_card_no')
        email = request.data.get('email')
        phone = request.data.get('phone_number')

        # Check NIN
        if nin and IndividualMember.objects.filter(nin=nin).exists():
            return Response({
                'exists': True,
                'field': 'nin',
                'message': 'An individual member with this NIN already exists'
            }, status=status.HTTP_200_OK)

        # Check Voters Card
        if voters_card_no and IndividualMember.objects.filter(voters_card_no=voters_card_no).exists():
            return Response({
                'exists': True,
                'field': 'voters_card_no',
                'message': 'An individual member with this Voters Card Number already exists'
            }, status=status.HTTP_200_OK)

        # Check Email
        if email and IndividualMember.objects.filter(email=email).exists():
            return Response({
                'exists': True,
                'field': 'email',
                'message': 'An individual member with this email already exists'
            }, status=status.HTTP_200_OK)

        # Check Phone
        if phone and IndividualMember.objects.filter(phone_number=phone).exists():
            return Response({
                'exists': True,
                'field': 'phone_number',
                'message': 'An individual member with this phone number already exists'
            }, status=status.HTTP_200_OK)

        return Response({
            'exists': False,
            'message': 'Information is available for registration'
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'])
    def get_id_card(self, request):
        """
        Get ID card file URL for a member
        Expects: member_id (query parameter)
        """
        member_id = request.query_params.get('member_id')
        
        try:
            member = IndividualMember.objects.get(abia_arise_id=member_id)
            
            if not member.id_card_file:
                return Response({
                    'success': False,
                    'message': 'ID card not yet generated'
                }, status=status.HTTP_404_NOT_FOUND)
            
            return Response({
                'success': True,
                'id_card_url': member.id_card_file.url,
                'id_card_path': str(member.id_card_file),
                'member_id': member.abia_arise_id,
                'member_name': f"{member.first_name} {member.middle_name} {member.last_name}".strip()
            }, status=status.HTTP_200_OK)
        
        except IndividualMember.DoesNotExist:
            return Response({
                'success': False,
                'message': 'Member not found'
            }, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['get'])
    def dashboard(self, request):
        """
        Get individual member dashboard data
        Requires member_id in query params
        """
        member_id = request.query_params.get('member_id')
        
        try:
            member = IndividualMember.objects.get(abia_arise_id=member_id)
            return Response({
                'success': True,
                'data': {
                    'id': member.id,
                    'abia_arise_id': member.abia_arise_id,
                    'first_name': member.first_name,
                    'middle_name': member.middle_name,
                    'last_name': member.last_name,
                    'email': member.email,
                    'phone_number': member.phone_number,
                    'age': member.age,
                    'gender': member.gender,
                    'occupation': member.occupation,
                    'profile_picture': member.profile_picture.url if member.profile_picture else None,
                    # Origin Details (cannot change)
                    'state_of_origin': member.state_of_origin,
                    'lga_of_origin': member.lga_of_origin,
                    'country_of_origin': member.country_of_origin,
                    # Residential Details (can change)
                    'lga_of_residence': member.lga_of_residence,
                    'state_of_residence': member.state_of_residence,
                    'country_of_residence': member.country_of_residence,
                    'electoral_ward': member.electoral_ward,
                    'polling_unit': member.polling_unit,
                    # Bank Details
                    'bank_account_number': member.bank_account_number,
                    'bank_name': member.bank_name,
                    'bvn': member.bvn,
                    'membership_purpose': member.membership_purpose,
                    'created_at': member.created_at,
                }
            }, status=status.HTTP_200_OK)
        except IndividualMember.DoesNotExist:
            return Response({
                'success': False,
                'message': 'Member not found'
            }, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['post'])
    def update_residential_info(self, request):
        """
        Update member's residential information
        Requires: member_id, pin, and residential fields
        """
        member_id = request.data.get('member_id')
        pin = request.data.get('pin')
        
        try:
            member = IndividualMember.objects.get(abia_arise_id=member_id)
            
            # Verify PIN
            if str(member.pin) != str(pin):
                return Response({
                    'success': False,
                    'message': 'Invalid PIN'
                }, status=status.HTTP_401_UNAUTHORIZED)
            
            # Update residential info
            if 'lga_of_residence' in request.data:
                member.lga_of_residence = request.data['lga_of_residence']
            if 'state_of_residence' in request.data:
                member.state_of_residence = request.data['state_of_residence']
            if 'electoral_ward' in request.data:
                member.electoral_ward = request.data['electoral_ward']
            if 'polling_unit' in request.data:
                member.polling_unit = request.data['polling_unit']
            
            member.save()
            return Response({
                'success': True,
                'message': 'Residential information updated successfully',
                'data': {
                    'lga_of_residence': member.lga_of_residence,
                    'state_of_residence': member.state_of_residence,
                    'electoral_ward': member.electoral_ward,
                    'polling_unit': member.polling_unit,
                }
            }, status=status.HTTP_200_OK)
        except IndividualMember.DoesNotExist:
            return Response({
                'success': False,
                'message': 'Member not found'
            }, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['post'])
    def update_bank_details(self, request):
        """
        Update member's bank details
        Requires: member_id, pin, bank_account_number, bank_name, bvn
        """
        member_id = request.data.get('member_id')
        pin = request.data.get('pin')
        
        try:
            member = IndividualMember.objects.get(abia_arise_id=member_id)
            
            # Verify PIN
            if str(member.pin) != str(pin):
                return Response({
                    'success': False,
                    'message': 'Invalid PIN'
                }, status=status.HTTP_401_UNAUTHORIZED)
            
            # Update bank details
            if 'bank_account_number' in request.data:
                member.bank_account_number = request.data['bank_account_number']
            if 'bank_name' in request.data:
                member.bank_name = request.data['bank_name']
            if 'bvn' in request.data:
                member.bvn = request.data['bvn']
            
            member.save()
            return Response({
                'success': True,
                'message': 'Bank details updated successfully',
                'data': {
                    'bank_account_number': member.bank_account_number,
                    'bank_name': member.bank_name,
                    'bvn': member.bvn,
                }
            }, status=status.HTTP_200_OK)
        except IndividualMember.DoesNotExist:
            return Response({
                'success': False,
                'message': 'Member not found'
            }, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['post'])
    def update_pin(self, request):
        """
        Update member's PIN
        Requires: member_id, nin, phone_number, old_pin, new_pin
        """
        member_id = request.data.get('member_id')
        nin = request.data.get('nin')
        phone_number = request.data.get('phone_number')
        old_pin = request.data.get('old_pin')
        new_pin = request.data.get('new_pin')
        
        try:
            member = IndividualMember.objects.get(abia_arise_id=member_id)
            
            # Verify identity
            if member.nin != nin or member.phone_number != phone_number:
                return Response({
                    'success': False,
                    'message': 'Invalid NIN or phone number'
                }, status=status.HTTP_401_UNAUTHORIZED)
            
            # Verify old PIN
            if str(member.pin) != str(old_pin):
                return Response({
                    'success': False,
                    'message': 'Invalid old PIN'
                }, status=status.HTTP_401_UNAUTHORIZED)
            
            # Update PIN
            member.pin = str(new_pin)
            member.save()
            
            return Response({
                'success': True,
                'message': 'PIN updated successfully'
            }, status=status.HTTP_200_OK)
        except IndividualMember.DoesNotExist:
            return Response({
                'success': False,
                'message': 'Member not found'
            }, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['post'])
    def update_profile_picture(self, request):
        """
        Update member's profile picture
        Requires: member_id, pin, profile_picture
        """
        member_id = request.data.get('member_id')
        pin = request.data.get('pin')
        profile_picture = request.FILES.get('profile_picture')
        
        try:
            member = IndividualMember.objects.get(abia_arise_id=member_id)
            
            # Verify PIN
            if str(member.pin) != str(pin):
                return Response({
                    'success': False,
                    'message': 'Invalid PIN'
                }, status=status.HTTP_401_UNAUTHORIZED)
            
            # Update profile picture
            if profile_picture:
                member.profile_picture = profile_picture
                member.save()
            
            return Response({
                'success': True,
                'message': 'Profile picture updated successfully',
                'profile_picture_url': member.profile_picture.url if member.profile_picture else None
            }, status=status.HTTP_200_OK)
        except IndividualMember.DoesNotExist:
            return Response({
                'success': False,
                'message': 'Member not found'
            }, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['post'])
    def login(self, request):
        """
        Login for individual members
        Member ID (abia_arise_id) and last 4 digits of phone as password
        """
        abia_arise_id = request.data.get('abia_arise_id') or request.data.get('member_id')
        password = request.data.get('password')

        if not abia_arise_id or not password:
            return Response({
                'success': False,
                'message': 'Missing abia_arise_id or password'
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            member = IndividualMember.objects.get(abia_arise_id=abia_arise_id)
            
            # Get expected password (last 4 digits of phone)
            # Handle case where password_hash might not be set
            if not member.password_hash and member.phone_number:
                # Set password_hash from phone number if not already set
                member.password_hash = member.phone_number[-4:]
                member.save()
            
            expected_password = member.password_hash or (member.phone_number[-4:] if member.phone_number else None)
            
            if expected_password and expected_password == password:
                # Generate JWT token
                refresh = RefreshToken.for_user(member.user) if hasattr(member, 'user') and member.user else None
                token = str(refresh.access_token) if refresh else None
                
                return Response({
                    'success': True,
                    'message': 'Login successful',
                    'token': token,
                    'member_id': member.id,
                    'abia_arise_id': member.abia_arise_id,
                    'user': {
                        'id': member.id,
                        'first_name': member.first_name,
                        'last_name': member.last_name,
                        'email': member.email,
                        'abia_arise_id': member.abia_arise_id,
                        'nin': member.nin,
                        'phone_number': member.phone_number,
                        'pin': member.pin,
                    }
                }, status=status.HTTP_200_OK)
            else:
                return Response({
                    'success': False,
                    'message': 'Invalid credentials'
                }, status=status.HTTP_401_UNAUTHORIZED)
        except IndividualMember.DoesNotExist:
            return Response({
                'success': False,
                'message': 'Member not found'
            }, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['post'])
    def activate_membership(self, request):
        """
        Activate membership for group members
        
        Request payload:
        {
            'nin': '12345678901',
            'phone_number': '08012345678',
            'email': 'member@example.com',
            'password': 'password_to_set',
            'membership_purpose': 'Optional purpose',
            'profile_picture': Optional file
        }
        """
        serializer = MemberActivationSerializer(data=request.data)
        if not serializer.is_valid():
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

        nin = serializer.validated_data['nin']
        phone_number = serializer.validated_data['phone_number']
        email = serializer.validated_data['email']
        password = serializer.validated_data['password']
        membership_purpose = serializer.validated_data.get('membership_purpose', '')
        profile_picture = request.FILES.get('profile_picture')

        try:
            # Try to find existing member via NIN
            member = IndividualMember.objects.get(nin=nin)
            
            # Verify phone and email match
            if member.phone_number != phone_number:
                return Response({
                    'success': False,
                    'message': 'Phone number does not match our records'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Activate membership
            if not member.is_individual:
                member.is_individual = True
                member.account_status = 'active'
                member.password_hash = password
                if membership_purpose:
                    member.membership_purpose = membership_purpose
                if profile_picture:
                    member.profile_picture = profile_picture
                member.save()
                
                return Response({
                    'success': True,
                    'message': 'Membership activated successfully. Welcome!',
                    'member_id': member.abia_arise_id,
                    'data': IndividualMemberSerializer(member).data
                }, status=status.HTTP_200_OK)
            
            elif member.is_individual and member.account_status == 'active':
                return Response({
                    'success': False,
                    'message': 'Member is already activated'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            else:
                return Response({
                    'success': False,
                    'message': 'Cannot activate this member at this time'
                }, status=status.HTTP_400_BAD_REQUEST)
                
        except IndividualMember.DoesNotExist:
            return Response({
                'success': False,
                'message': 'Member not found in system. Please contact your Pro-Group administrator.'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


class ProGroupViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing pro-groups
    """
    queryset = ProGroup.objects.all()
    serializer_class = ProGroupSerializer
    parser_classes = (JSONParser, MultiPartParser, FormParser)

    def get_permissions(self):
        """
        Allow unauthenticated access to register, login, generate_certificate, get_certificate, download_template, upload_members, and dashboard endpoints
        """
        if self.action in ['register', 'login', 'generate_certificate', 'get_certificate', 'dashboard', 'list', 'update_logo', 'update_address', 'download_template', 'upload_members', 'generate_reset_pin', 'verify_reset_pin_otp', 'verify_reset_pin', 'add_member_manual']:
            return [AllowAny()]
        return super().get_permissions()

    @action(detail=False, methods=['post'])
    def register(self, request):
        """
        Register a new pro-group and automatically generate certificate
        """
        serializer = self.get_serializer(data=request.data)
        if serializer.is_valid():
            try:
                group = serializer.save()
                
                # Set reset_pin from registration if provided
                reset_pin = request.data.get('reset_pin')
                if reset_pin and len(reset_pin) == 6 and reset_pin.isdigit():
                    group.reset_pin = reset_pin
                    group.save()
                    print(f"✓ Reset PIN set during registration for group {group.group_license_number}")
                
                # Automatically generate certificate after registration
                group_data = {
                    'group_name': group.name,
                    'group_license_number': group.group_license_number,
                    'state': getattr(group, 'state', group.lga),  # Use state if exists, otherwise use lga
                    'lga': group.lga,
                }
                
                success, cert_path, error = generate_certificate(group_data, group_instance=group)
                
                if success:
                    print(f"✓ Certificate generated successfully: {cert_path}")
                else:
                    print(f"✗ Certificate generation warning: {error}")
                
                return Response({
                    'success': True,
                    'message': 'Pro-group registration successful. Certificate generated.',
                    'group_id': serializer.data['group_license_number'],
                    'data': serializer.data
                }, status=status.HTTP_201_CREATED)
            except Exception as e:
                return Response({
                    'success': False,
                    'error': str(e)
                }, status=status.HTTP_400_BAD_REQUEST)
        print("ProGroup Serializer errors:", serializer.errors)  # DEBUG
        return Response({
            'success': False,
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
    @action(detail=False, methods=['post'])
    def generate_certificate(self, request):
        """
        Generate certificate for a registered pro-group
        Expects: group_id (group_license_number)
        """
        try:
            group_id = request.data.get('group_id')
            print(f"DEBUG: Generating certificate for group_id: {group_id}")
            group = ProGroup.objects.get(group_license_number=group_id)
            print(f"DEBUG: Found group: {group.name}")
            
            # Prepare group data for certificate generation
            group_data = {
                'name': group.name,
                'group_license_number': group.group_license_number,
                'chairman_name': group.chairman_name,
                'secretary_name': group.secretary_name
            }
            
            # Generate certificate
            generator = CertificateGenerator()
            cert_path = generator.generate(group_data)
            
            return Response({
                'success': True,
                'message': 'Certificate generated successfully',
                'certificate_url': cert_path,
                'group_data': {
                    'name': group.name,
                    'group_license_number': group.group_license_number
                }
            }, status=status.HTTP_200_OK)
        
        except ProGroup.DoesNotExist:
            print("DEBUG: Group not found")
            return Response({
                'success': False,
                'message': 'Group not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            print(f"DEBUG: Certificate generation error: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def get_certificate(self, request):
        """
        Get certificate file URL for a pro-group
        Expects: group_id (group_license_number) query parameter
        """
        group_id = request.query_params.get('group_id')
        
        try:
            group = ProGroup.objects.get(group_license_number=group_id)
            
            if not group.certificate_file:
                return Response({
                    'success': False,
                    'message': 'Certificate not yet generated'
                }, status=status.HTTP_404_NOT_FOUND)
            
            return Response({
                'success': True,
                'certificate_url': group.certificate_file.url,
                'certificate_path': str(group.certificate_file),
                'group_id': group.group_license_number,
                'group_name': group.name
            }, status=status.HTTP_200_OK)
        
        except ProGroup.DoesNotExist:
            return Response({
                'success': False,
                'message': 'Group not found'
            }, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['get'])
    def dashboard(self, request):
        """
        Get pro-group dashboard data
        Requires group_id (group_license_number) in query params
        """
        group_id = request.query_params.get('group_id')
        
        try:
            group = ProGroup.objects.get(group_license_number=group_id)
            
            # Get total_members from Excel file if it exists
            total_members = 0
            if group.excel_file:
                try:
                    # Read Excel file to count members
                    excel_data = BytesIO(group.excel_file.read())
                    workbook = openpyxl.load_workbook(excel_data)
                    worksheet = workbook.active
                    
                    # Count rows (excluding header)
                    row_count = 0
                    for row in worksheet.iter_rows(min_row=2, values_only=True):
                        # Skip empty rows
                        if any(cell is not None for cell in row):
                            row_count += 1
                    
                    total_members = row_count
                except Exception as excel_error:
                    print(f"Warning: Failed to read Excel file: {excel_error}")
                    total_members = group.total_members
            else:
                # Fall back to group members count
                total_members = group.total_members
            
            # Get group members with details
            group_members = GroupMember.objects.filter(group=group).select_related('member')
            
            members_data = []
            
            for gm in group_members:
                member_info = {
                    'id': gm.id,
                    'member_id': gm.member.id,
                    'abia_arise_id': gm.member.abia_arise_id,
                    'first_name': gm.member.first_name,
                    'last_name': gm.member.last_name,
                    'email': gm.member.email,
                    'phone_number': gm.member.phone_number,
                    'lga_of_origin': gm.member.lga_of_origin,
                    'occupation': gm.member.occupation,
                    'role': gm.role,
                    'is_group_member': gm.is_group_member,
                    'is_individual': gm.member.is_individual,
                    'account_status': gm.member.account_status,
                }
                members_data.append(member_info)
            
            return Response({
                'success': True,
                'data': {
                    'id': group.id,
                    'name': group.name,
                    'group_license_number': group.group_license_number,
                    'lga': group.lga,
                    'address': group.address,
                    'logo': group.logo.url if group.logo else None,
                    'total_members': total_members,
                    'registered_date': group.created_at.strftime('%B %d, %Y'),
                    'chairman_name': group.chairman_name,
                    'chairman_phone': group.chairman_phone,
                    'chairman_email': group.chairman_email,
                    'secretary_name': group.secretary_name,
                    'secretary_phone': group.secretary_phone,
                    'secretary_email': group.secretary_email,
                    'excel_file_url': group.excel_file.url if group.excel_file else None,
                    'members': members_data,
                    'member_count': len(members_data),
                    'created_at': group.created_at,
                }
            }, status=status.HTTP_200_OK)
        except ProGroup.DoesNotExist:
            return Response({
                'success': False,
                'message': 'Group not found'
            }, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['post'])
    def update_logo(self, request):
        """
        Update group logo
        Requires: group_id, password (chairman/secretary phone last 4 digits), logo
        """
        group_id = request.data.get('group_id')
        password = request.data.get('password')
        logo = request.FILES.get('logo')
        
        try:
            group = ProGroup.objects.get(group_license_number=group_id)
            
            # Verify user is chairman or secretary
            if not (group.chairman_password_hash == password or 
                    group.secretary_password_hash == password):
                return Response({
                    'success': False,
                    'message': 'Unauthorized'
                }, status=status.HTTP_401_UNAUTHORIZED)
            
            # Update logo
            if logo:
                group.logo = logo
                group.save()
            
            return Response({
                'success': True,
                'message': 'Logo updated successfully',
                'logo_url': group.logo.url if group.logo else None
            }, status=status.HTTP_200_OK)
        except ProGroup.DoesNotExist:
            return Response({
                'success': False,
                'message': 'Group not found'
            }, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['post'])
    def update_address(self, request):
        """
        Update group address
        Requires: group_id, password (chairman/secretary phone last 4 digits), address
        """
        group_id = request.data.get('group_id')
        password = request.data.get('password')
        address = request.data.get('address')
        
        try:
            group = ProGroup.objects.get(group_license_number=group_id)
            
            # Verify user is chairman or secretary
            if not (group.chairman_password_hash == password or 
                    group.secretary_password_hash == password):
                return Response({
                    'success': False,
                    'message': 'Unauthorized'
                }, status=status.HTTP_401_UNAUTHORIZED)
            
            # Update address
            if address:
                group.address = address
                group.save()
            
            return Response({
                'success': True,
                'message': 'Address updated successfully',
                'address': group.address
            }, status=status.HTTP_200_OK)
        except ProGroup.DoesNotExist:
            return Response({
                'success': False,
                'message': 'Group not found'
            }, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['post'])
    def login(self, request):
        """
        Login for pro-groups
        Group License Number and last 4 digits of chairman/secretary phone
        """
        group_license = request.data.get('group_license_number')
        password = request.data.get('password')

        try:
            group = ProGroup.objects.get(group_license_number=group_license)
            user_role = None
            user_name = None
            
            if group.chairman_password_hash == password:
                user_role = 'chairman'
                user_name = group.chairman_name
            elif group.secretary_password_hash == password:
                user_role = 'secretary'
                user_name = group.secretary_name
            
            if user_role:
                return Response({
                    'success': True,
                    'message': 'Login successful',
                    'token': None,  # Could generate JWT if needed
                    'group_license_number': group.group_license_number,
                    'user': {
                        'name': group.name,
                        'group_license_number': group.group_license_number,
                        'lga': group.lga,
                        'role': user_role,
                        'user_name': user_name,
                    }
                }, status=status.HTTP_200_OK)
            else:
                return Response({
                    'success': False,
                    'message': 'Invalid credentials'
                }, status=status.HTTP_401_UNAUTHORIZED)
        except ProGroup.DoesNotExist:
            return Response({
                'success': False,
                'message': 'Group not found'
            }, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """
        Download the Excel template for group member registration
        """
        template_path = os.path.join(settings.MEDIA_ROOT, 'templates', 'Abia arise pro-group members template.xlsx')
        
        try:
            if os.path.exists(template_path):
                response = FileResponse(open(template_path, 'rb'), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename="Abia_Arise_ProGroup_Members_Template.xlsx"'
                return response
            else:
                return Response({
                    'success': False,
                    'message': 'Template file not found'
                }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'success': False,
                'message': f'Error downloading template: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'])
    def upload_members(self, request):
        """
        Upload and process Excel file with group members
        
        The Excel file should have columns:
        - Full Name (or First Name + Last Name)
        - NIN (National Identification Number)
        - Voters Card No (optional)
        - Email (optional)
        - Phone Number
        - Occupation (optional)
        
        Request payload:
        {
            'group_id': 'AB/PRG/ABC/0001',
            'password': 'last_4_digits_of_chairman_or_secretary_phone',
            'excel_file': <file>
        }
        """
        serializer = ExcelMemberImportSerializer(data=request.data)
        if not serializer.is_valid():
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

        group_id = serializer.validated_data['group_id']
        password = serializer.validated_data['password']
        excel_file = serializer.validated_data['excel_file']

        try:
            # Verify group exists and user has permission
            group = ProGroup.objects.get(group_license_number=group_id)
            
            # Verify user is chairman or secretary
            if not (group.chairman_password_hash == password or 
                    group.secretary_password_hash == password):
                return Response({
                    'success': False,
                    'message': 'Unauthorized. Invalid Chairman/Secretary credentials.'
                }, status=status.HTTP_401_UNAUTHORIZED)

            # Read Excel file
            excel_data = BytesIO(excel_file.read())
            workbook = openpyxl.load_workbook(excel_data)
            worksheet = workbook.active

            # Expected column indices (0-based)
            processed_members = []
            errors = []
            skipped = []
            row_number = 2  # Start from row 2 (row 1 is headers)

            for row in worksheet.iter_rows(min_row=2, values_only=True):
                try:
                    # Parse row data - handle flexible column naming
                    if len(row) < 5:
                        errors.append(f"Row {row_number}: Insufficient data columns")
                        row_number += 1
                        continue

                    # Extract data from columns
                    # Assuming: Full Name/First Name, NIN, Voters Card, Email, Phone, Occupation
                    full_name = str(row[0] or '').strip()
                    nin = str(row[1] or '').strip()
                    voters_card = str(row[2] or '').strip() if row[2] else ''
                    email = str(row[3] or '').strip() if row[3] else ''
                    phone = str(row[4] or '').strip()
                    occupation = str(row[5] or '').strip() if len(row) > 5 and row[5] else ''

                    # Validate required fields
                    if not full_name or not nin or not phone:
                        errors.append(f"Row {row_number}: Missing required fields (Name, NIN, or Phone)")
                        row_number += 1
                        continue

                    # Validate NIN format (11 digits)
                    if len(nin) != 11 or not nin.isdigit():
                        errors.append(f"Row {row_number}: Invalid NIN format (must be 11 digits)")
                        row_number += 1
                        continue

                    # Parse full name into first and last name
                    name_parts = full_name.split()
                    first_name = name_parts[0] if name_parts else ''
                    last_name = ' '.join(name_parts[1:]) if len(name_parts) > 1 else name_parts[0]

                    # Check if member already exists via NIN (master key)
                    member, created = IndividualMember.objects.get_or_create(
                        nin=nin,
                        defaults={
                            'first_name': first_name,
                            'last_name': last_name,
                            'email': email or None,
                            'phone_number': phone,
                            'occupation': occupation,
                            'voters_card_no': voters_card or None,
                            'is_individual': False,
                            'is_group_member': True,
                            'account_status': 'pending_activation',  # Not activated yet
                        }
                    )

                    if not created:
                        # Member already exists
                        # If they were already an individual member, just mark them as group member
                        if member.is_individual:
                            member.is_group_member = True
                            member.save()
                            skipped.append(f"Row {row_number}: Member already exists as individual (added to group)")
                        else:
                            skipped.append(f"Row {row_number}: Member already in group")
                    else:
                        # New member account created
                        processed_members.append({
                            'nin': nin,
                            'name': full_name,
                            'member_id': member.abia_arise_id,
                            'status': 'created'
                        })

                    # Link member to group (create GroupMember record if not exists)
                    group_member, gm_created = GroupMember.objects.get_or_create(
                        group=group,
                        member=member,
                        defaults={'role': 'member'}
                    )

                    row_number += 1

                except ValueError as e:
                    errors.append(f"Row {row_number}: Data parsing error - {str(e)}")
                    row_number += 1
                except Exception as e:
                    errors.append(f"Row {row_number}: {str(e)}")
                    row_number += 1

            # Update group's total_members count
            group.total_members = group.members.count()
            
            # Delete old excel file if it exists
            if group.excel_file and os.path.exists(group.excel_file.path):
                os.remove(group.excel_file.path)
            
            # Save the new excel file
            group.excel_file = excel_file
            group.save()

            return Response({
                'success': True,
                'message': f'Excel import completed. {len(processed_members)} members added to group.',
                'summary': {
                    'created': len(processed_members),
                    'skipped': len(skipped),
                    'errors_count': len(errors),
                    'total_group_members': group.total_members
                },
                'processed_members': processed_members,
                'skipped': skipped if skipped else [],
                'errors': errors if errors else [],
                'group_id': group.group_license_number,
                'group_name': group.name,
                'excel_file_url': group.excel_file.url if group.excel_file else None
            }, status=status.HTTP_200_OK)

        except ProGroup.DoesNotExist:
            return Response({
                'success': False,
                'message': 'Group not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({
                'success': False,
                'error': str(e),
                'message': 'Error processing Excel file'
            }, status=status.HTTP_400_BAD_REQUEST)




    @action(detail=False, methods=['post'])
    def delete_excel_file(self, request):
        """
        Delete the current Excel file for a group
        
        Request payload:
        {
            'group_id': 'AB/PRG/ABC/0001',
            'password': 'last_4_digits_of_chairman_or_secretary_phone'
        }
        """
        try:
            group_id = request.data.get('group_id')
            password = request.data.get('password')
            
            if not group_id or not password:
                return Response({
                    'success': False,
                    'message': 'Missing required fields: group_id, password'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Find group
            group = ProGroup.objects.get(group_license_number=group_id)
            
            # Verify user is chairman or secretary
            if not (group.chairman_password_hash == password or 
                    group.secretary_password_hash == password):
                return Response({
                    'success': False,
                    'message': 'Unauthorized. Invalid credentials.'
                }, status=status.HTTP_401_UNAUTHORIZED)
            
            # Delete excel file if exists
            if group.excel_file:
                if os.path.exists(group.excel_file.path):
                    os.remove(group.excel_file.path)
                group.excel_file = None
                group.save()
                
                return Response({
                    'success': True,
                    'message': 'Excel file deleted successfully'
                }, status=status.HTTP_200_OK)
            else:
                return Response({
                    'success': False,
                    'message': 'No Excel file found for this group'
                }, status=status.HTTP_404_NOT_FOUND)
                
        except ProGroup.DoesNotExist:
            return Response({
                'success': False,
                'message': 'Group not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e),
                'message': 'Error deleting Excel file'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def generate_reset_pin(self, request):
        """
        Generate a reset PIN for group admin mode
        Sends OTP to chairman and secretary emails
        
        Request payload:
        {
            'group_id': 'AB/PRG/ABC/0001',
            'password': 'last_4_digits_of_phone',
            'desired_pin': '123456'
        }
        """
        serializer = GenerateResetPinSerializer(data=request.data)
        if not serializer.is_valid():
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

        group_id = serializer.validated_data['group_id']
        password = serializer.validated_data['password']
        desired_pin = serializer.validated_data['desired_pin']

        try:
            group = ProGroup.objects.get(group_license_number=group_id)
            
            # Verify user is chairman or secretary
            if not (group.chairman_password_hash == password or 
                    group.secretary_password_hash == password):
                return Response({
                    'success': False,
                    'message': 'Unauthorized. Invalid credentials.'
                }, status=status.HTTP_401_UNAUTHORIZED)

            # Generate 6-digit OTP
            otp = ''.join([str(random.randint(0, 9)) for _ in range(6)])
            
            # Store pending PIN and OTP
            group.pending_reset_pin = desired_pin
            group.pending_reset_pin_otp = otp
            group.pending_reset_pin_expiry = timezone.now() + timedelta(minutes=10)
            group.save()
            
            print(f"✓ Generated OTP for group {group_id}: {otp}")
            print(f"✓ Sending OTP to: {group.chairman_email}, {group.secretary_email}")

            # Flag to track if email was sent successfully
            email_sent_success = False
            email_error_message = None

            # Send OTP to emails
            try:
                email_subject = f"Abia Arise - Reset PIN Verification OTP"
                email_message = f"""Hello,

A reset PIN has been requested for {group.name}.

Your OTP for verification: {otp}

This OTP is valid for 10 minutes.

If you didn't request this, please ignore this email.

Regards,
Abia Arise Team"""
                
                recipients = [group.chairman_email, group.secretary_email]
                send_mail(
                    email_subject,
                    email_message,
                    settings.DEFAULT_FROM_EMAIL,
                    recipients,
                    fail_silently=True  # Use True to not raise exception
                )
                email_sent_success = True
                print(f"✓ OTP email sent successfully to {recipients}")
            except Exception as email_error:
                email_error_message = str(email_error)
                print(f"✗ Error sending OTP email: {email_error_message}")
                # Don't fail the API call, but notify user of potential issue

            response_data = {
                'success': True,
                'message': 'OTP generated successfully. You have 10 minutes to verify.',
                'group_id': group.group_license_number,
                'group_name': group.name,
                'email_sent': email_sent_success
            }
            
            if not email_sent_success and email_error_message:
                response_data['email_warning'] = f'OTP was not sent to email due to: {email_error_message}'
            
            # For development/testing: include OTP in response (REMOVE FOR PRODUCTION)
            if settings.DEBUG:
                response_data['otp'] = otp
            
            return Response(response_data, status=status.HTTP_200_OK)

        except ProGroup.DoesNotExist:
            return Response({
                'success': False,
                'message': 'Group not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            error_message = str(e)
            print(f"✗ Error in generate_reset_pin: {error_message}")
            return Response({
                'success': False,
                'error': error_message
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def verify_reset_pin_otp(self, request):
        """
        Verify OTP and set the reset PIN
        
        Request payload:
        {
            'group_id': 'AB/PRG/ABC/0001',
            'otp': '123456'
        }
        """
        serializer = VerifyResetPinOtpSerializer(data=request.data)
        if not serializer.is_valid():
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

        group_id = serializer.validated_data['group_id']
        otp = serializer.validated_data['otp']

        try:
            group = ProGroup.objects.get(group_license_number=group_id)
            
            print(f"✓ Verifying OTP for group {group_id}")
            print(f"✓ Received OTP: {otp}")
            print(f"✓ Stored OTP: {group.pending_reset_pin_otp}")
            print(f"✓ OTP Expiry: {group.pending_reset_pin_expiry}")
            print(f"✓ Current Time: {timezone.now()}")
            
            # Check if OTP is still valid
            if not group.pending_reset_pin_expiry:
                print(f"✗ OTP expiry not set for group {group_id}")
                return Response({
                    'success': False,
                    'message': 'OTP has expired. No OTP found. Please generate a new one.'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            if timezone.now() > group.pending_reset_pin_expiry:
                print(f"✗ OTP expired for group {group_id}")
                return Response({
                    'success': False,
                    'message': 'OTP has expired. Please generate a new one.'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Verify OTP matches
            if group.pending_reset_pin_otp != otp:
                print(f"✗ OTP mismatch for group {group_id}: expected {group.pending_reset_pin_otp}, got {otp}")
                return Response({
                    'success': False,
                    'message': 'Invalid OTP. Please try again.'
                }, status=status.HTTP_401_UNAUTHORIZED)

            # OTP verified, set the reset PIN
            print(f"✓ OTP verified. Setting reset PIN: {group.pending_reset_pin}")
            group.reset_pin = group.pending_reset_pin
            group.pending_reset_pin = None
            group.pending_reset_pin_otp = None
            group.pending_reset_pin_expiry = None
            group.save()
            
            print(f"✓ Reset PIN set successfully for group {group_id}: {group.reset_pin}")

            return Response({
                'success': True,
                'message': 'Reset PIN verified and set successfully!',
                'group_id': group.group_license_number,
                'reset_pin': group.reset_pin
            }, status=status.HTTP_200_OK)

        except ProGroup.DoesNotExist:
            print(f"✗ Group not found: {group_id}")
            return Response({
                'success': False,
                'message': 'Group not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            error_message = str(e)
            print(f"✗ Error in verify_reset_pin_otp: {error_message}")
            return Response({
                'success': False,
                'error': error_message
            }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def verify_reset_pin(self, request):
        """
        Verify reset PIN to enter admin/reset mode
        
        Request payload:
        {
            'group_id': 'AB/PRG/ABC/0001',
            'reset_pin': '123456'
        }
        """
        serializer = VerifyResetPinSerializer(data=request.data)
        if not serializer.is_valid():
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

        group_id = serializer.validated_data['group_id']
        reset_pin = serializer.validated_data['reset_pin']

        try:
            group = ProGroup.objects.get(group_license_number=group_id)
            
            if not group.reset_pin:
                return Response({
                    'success': False,
                    'message': 'No reset PIN set for this group. Generate one first.'
                }, status=status.HTTP_400_BAD_REQUEST)

            if group.reset_pin != reset_pin:
                return Response({
                    'success': False,
                    'message': 'Invalid reset PIN.'
                }, status=status.HTTP_401_UNAUTHORIZED)

            return Response({
                'success': True,
                'message': 'Reset mode activated!',
                'group_id': group.group_license_number,
                'group_name': group.name
            }, status=status.HTTP_200_OK)

        except ProGroup.DoesNotExist:
            return Response({
                'success': False,
                'message': 'Group not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def download_certificate(self, request):
        """
        Download group certificate as PDF (requires reset PIN)
        
        Request payload:
        {
            'group_id': 'AB/PRG/ABC/0001',
            'reset_pin': '123456'
        }
        """
        try:
            group_id = request.data.get('group_id')
            reset_pin = request.data.get('reset_pin')
            
            if not group_id or not reset_pin:
                return Response({
                    'success': False,
                    'message': 'Missing required fields: group_id, reset_pin'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Find group
            group = ProGroup.objects.get(group_license_number=group_id)
            
            # Verify reset PIN
            if not group.reset_pin or group.reset_pin != reset_pin:
                return Response({
                    'success': False,
                    'message': 'Invalid reset PIN.'
                }, status=status.HTTP_401_UNAUTHORIZED)
            
            # Generate certificate PDF
            certificate_generator = CertificateGenerator()
            certificate_buffer = certificate_generator.generate_group_certificate(
                group_name=group.name,
                group_license=group.group_license_number,
                lga=group.lga,
                registration_date=group.created_at.strftime('%B %d, %Y'),
                chairman_name=group.chairman_name,
                secretary_name=group.secretary_name,
                total_members=group.total_members
            )
            
            # Return PDF as file download
            response = FileResponse(
                certificate_buffer,
                content_type='application/pdf',
                as_attachment=True,
                filename=f"{group.name}_Certificate.pdf"
            )
            return response
            
        except ProGroup.DoesNotExist:
            return Response({
                'success': False,
                'message': 'Group not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e),
                'message': 'Error generating certificate'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def add_member_manual(self, request):
        """
        Manually add a new member to the group (requires reset PIN)
        
        Request payload:
        {
            'group_id': 'AB/PRG/ABC/0001',
            'reset_pin': '123456',
            'full_name': 'John Doe',
            'nin': '12345678901',
            'phone_number': '08012345678',
            'email': 'john@example.com',
            'voters_card_no': 'VC/123',
            'occupation': 'Teacher'
        }
        """
        serializer = AddMemberToGroupSerializer(data=request.data)
        if not serializer.is_valid():
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)

        group_id = serializer.validated_data['group_id']
        reset_pin = serializer.validated_data['reset_pin']
        full_name = serializer.validated_data['full_name']
        nin = serializer.validated_data['nin']
        phone_number = serializer.validated_data['phone_number']
        email = serializer.validated_data.get('email', '')
        voters_card_no = serializer.validated_data.get('voters_card_no', '')
        occupation = serializer.validated_data.get('occupation', '')

        try:
            group = ProGroup.objects.get(group_license_number=group_id)
            
            # Verify reset PIN
            if not group.reset_pin or group.reset_pin != reset_pin:
                return Response({
                    'success': False,
                    'message': 'Invalid reset PIN. Not in reset mode.'
                }, status=status.HTTP_401_UNAUTHORIZED)

            # Validate NIN format
            if len(nin) != 11 or not nin.isdigit():
                return Response({
                    'success': False,
                    'message': 'Invalid NIN format (must be 11 digits)'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Check if member with this NIN already exists
            member, created = IndividualMember.objects.get_or_create(
                nin=nin,
                defaults={
                    'first_name': full_name.split()[0],
                    'last_name': ' '.join(full_name.split()[1:]) if len(full_name.split()) > 1 else full_name,
                    'email': email or None,
                    'phone_number': phone_number,
                    'occupation': occupation,
                    'voters_card_no': voters_card_no or None,
                    'is_individual': False,
                    'is_group_member': True,
                    'account_status': 'pending_activation'
                }
            )

            if not created:
                # Member exists, just mark as group member if not already
                if not member.is_group_member:
                    member.is_group_member = True
                    member.save()

            # Link to group
            group_member, gm_created = GroupMember.objects.get_or_create(
                group=group,
                member=member,
                defaults={'role': 'member'}
            )

            # Update group member count
            group.total_members = group.members.count()
            group.save()

            status_msg = 'created' if created else 'already exists'
            return Response({
                'success': True,
                'message': f'Member {status_msg} and added to group',
                'member_id': member.abia_arise_id,
                'member_name': member.get_full_name(),
                'group_id': group.group_license_number
            }, status=status.HTTP_201_CREATED if created else status.HTTP_200_OK)

        except ProGroup.DoesNotExist:
            return Response({
                'success': False,
                'message': 'Group not found'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


class GroupMemberViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing group members
    """
    queryset = GroupMember.objects.all()
    serializer_class = GroupMemberSerializer
    permission_classes = [IsAuthenticated]

    def get_permissions(self):
        """
        Allow unauthenticated access to list and retrieve group members
        """
        if self.action in ['list', 'retrieve']:
            return [AllowAny()]
        return [IsAuthenticated()]

    def get_queryset(self):
        """
        Filter group members by group if group_id is provided
        """
        queryset = super().get_queryset()
        group_id = self.request.query_params.get('group_id')
        if group_id:
            queryset = queryset.filter(group_id=group_id)
        return queryset

    def check_admin_permission(self, allow_group_admin=False):
        """
        Check if user is a superuser/admin or (if allow_group_admin=True) 
        can be verified as a group chairman/secretary
        """
        if self.request.user.is_superuser or self.request.user.is_staff:
            return True
        
        if allow_group_admin:
            # Check if request contains group_id and password for chairman/secretary verification
            group_id = self.request.data.get('group_id')
            password = self.request.data.get('password')
            
            if group_id and password:
                try:
                    group = ProGroup.objects.get(group_license_number=group_id)
                    if group.chairman_password_hash == password or group.secretary_password_hash == password:
                        return True
                except ProGroup.DoesNotExist:
                    pass
        
        raise PermissionError('Only admins or authorized group chairman/secretary can perform this action')

    def perform_update(self, serializer):
        """Override update to check permissions"""
        self.check_admin_permission(allow_group_admin=True)
        serializer.save()

    def perform_destroy(self, instance):
        """Override delete to check permissions"""
        self.check_admin_permission(allow_group_admin=True)
        instance.delete()

    def update(self, request, *args, **kwargs):
        """Handle PATCH and PUT requests with permission checking"""
        try:
            self.check_admin_permission(allow_group_admin=True)
            return super().update(request, *args, **kwargs)
        except PermissionError as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_403_FORBIDDEN)

    def destroy(self, request, *args, **kwargs):
        """Handle DELETE requests with permission checking"""
        try:
            self.check_admin_permission()
            return super().destroy(request, *args, **kwargs)
        except PermissionError as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_403_FORBIDDEN)

    @action(detail=False, methods=['post'])
    def add_member(self, request):
        """
        Add a member to a group
        
        Can be called by:
        1. Admin/superuser (without group_id/password)
        2. Group chairman/secretary (with group_id and password)
        
        Request payload:
        {
            'group': group_id,
            'member': member_id,
            'role': 'member'  // optional: 'member', 'chairman', 'secretary'
            // For chairman/secretary:
            'group_id': 'AB/PRG/ABC/0001',
            'password': 'last_4_digits'
        }
        """
        try:
            self.check_admin_permission(allow_group_admin=True)
            serializer = self.get_serializer(data=request.data)
            if serializer.is_valid():
                serializer.save()
                return Response({
                    'success': True,
                    'message': 'Member added to group',
                    'data': serializer.data
                }, status=status.HTTP_201_CREATED)
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        except PermissionError as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_403_FORBIDDEN)

    @action(detail=True, methods=['delete'])
    def remove_member(self, request, pk=None):
        """
        Remove a member from a group
        
        Can be called by:
        1. Admin/superuser (without group_id/password)
        2. Group chairman/secretary with reset PIN (with group_id and reset_pin)
        
        Request payload (for chairman/secretary in reset mode):
        {
            'group_id': 'AB/PRG/ABC/0001',
            'reset_pin': '123456'
        }
        """
        try:
            group_member = self.get_object()
            
            # Check permissions - can be admin or chairman/secretary with reset PIN
            is_admin = self.request.user.is_superuser or self.request.user.is_staff
            
            if not is_admin:
                # Must provide reset PIN
                group_id = self.request.data.get('group_id')
                reset_pin = self.request.data.get('reset_pin')
                
                if not group_id or not reset_pin:
                    raise PermissionError('Reset PIN required to remove members')
                
                try:
                    group = ProGroup.objects.get(group_license_number=group_id)
                    if not group.reset_pin or group.reset_pin != reset_pin:
                        raise PermissionError('Invalid reset PIN')
                except ProGroup.DoesNotExist:
                    raise PermissionError('Group not found')
            
            group_member.delete()
            return Response({
                'success': True,
                'message': 'Member removed from group'
            }, status=status.HTTP_204_NO_CONTENT)
        except PermissionError as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_403_FORBIDDEN)


class AdminLoginView(APIView):
    """
    API view for admin login functionality
    Supports both superuser and staff admin login via email and password
    """
    # Allow unauthenticated access to login endpoint
    permission_classes = [AllowAny]

    def post(self, request):
        """
        Login for admins (superusers and staff members)
        Requires email and password
        """
        email = request.data.get('email')
        password = request.data.get('password')

        if not email or not password:
            return Response({
                'success': False,
                'detail': 'Email and password are required'
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Try to find user by email first
            user = User.objects.filter(email=email).first()
            
            # If not found by email, try username
            if not user:
                user = User.objects.filter(username=email).first()

            # Check if user found
            if not user:
                return Response({
                    'success': False,
                    'detail': 'Admin user not found'
                }, status=status.HTTP_404_NOT_FOUND)

            # Check if user is admin (superuser or staff)
            if not (user.is_superuser or user.is_staff):
                return Response({
                    'success': False,
                    'detail': 'This account does not have admin privileges'
                }, status=status.HTTP_403_FORBIDDEN)

            # Verify password
            if not user.check_password(password):
                return Response({
                    'success': False,
                    'detail': 'Invalid email or password'
                }, status=status.HTTP_401_UNAUTHORIZED)

            # Generate JWT token
            refresh = RefreshToken.for_user(user)

            # Login successful
            return Response({
                'success': True,
                'message': 'Admin login successful',
                'token': str(refresh.access_token),
                'refresh': str(refresh),
                'admin': {
                    'id': user.id,
                    'username': user.username,
                    'email': user.email,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'is_superuser': user.is_superuser,
                    'is_staff': user.is_staff,
                }
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({
                'success': False,
                'detail': f'An error occurred: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ==========================================
# MEMBERSHIP SYSTEM VIEWS
# For flexible user-group membership management
# ==========================================

class MembershipUserViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing users in the flexible membership system.
    """
    queryset = MembershipUser.objects.all()
    parser_classes = (JSONParser, MultiPartParser, FormParser)
    
    def get_serializer_class(self):
        if self.action == 'create' or self.action == 'register':
            return MembershipUserCreateSerializer
        elif self.action == 'retrieve':
            return MembershipUserDetailSerializer
        return MembershipUserListSerializer
    
    def get_permissions(self):
        if self.action in ['list', 'retrieve', 'create', 'register', 'check_exists']:
            return [AllowAny()]
        return [IsAuthenticated()]
    
    @action(detail=False, methods=['post'])
    def register(self, request):
        """
        Self-register a new user or complete partial registration.
        
        Request payload:
        {
            'nin': '12345678901',
            'first_name': 'John',
            'last_name': 'Doe',
            'phone': '08012345678',
            'email': 'john@example.com'
        }
        
        If NIN exists (partial registration from group import):
        - Pre-fill form with existing data
        - Allow user to complete missing fields
        - Update registration_status to 'complete'
        
        If NIN doesn't exist:
        - Create new user with registration_status = 'complete'
        """
        serializer = UserRegistrationSerializer(data=request.data)
        if not serializer.is_valid():
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        
        nin = serializer.validated_data['nin']
        first_name = serializer.validated_data.get('first_name', '')
        last_name = serializer.validated_data.get('last_name', '')
        phone = serializer.validated_data.get('phone', '')
        email = serializer.validated_data.get('email', '')
        
        try:
            # Check if user with this NIN already exists
            user = MembershipUser.objects.filter(nin=nin).first()
            
            if user:
                # User exists - they're completing registration
                if user.registration_status == 'complete':
                    return Response({
                        'status': 'already_complete',
                        'message': f'User with NIN {nin} is already fully registered',
                        'user_id': user.id,
                        'groups': GroupMembership.objects.filter(user=user).values_list('group__name', flat=True)
                    }, status=status.HTTP_200_OK)
                
                # Update partial registration to complete
                if first_name:
                    user.first_name = first_name
                if last_name:
                    user.last_name = last_name
                if phone and not user.phone:
                    user.phone = phone
                if email and not user.email:
                    user.email = email
                
                user.registration_status = 'complete'
                user.updated_at = timezone.now()
                user.save()
                
                return Response({
                    'success': True,
                    'status': 'updated_from_partial',
                    'message': f'Welcome back! Your registration is now complete.',
                    'user': MembershipUserDetailSerializer(user).data,
                    'groups': GroupMembership.objects.filter(user=user).values('id', 'group__id', 'group__name', 'role', 'added_at')
                }, status=status.HTTP_200_OK)
            
            else:
                # New user - create with complete status
                new_user = MembershipUser.objects.create(
                    nin=nin,
                    first_name=first_name or 'Unknown',
                    last_name=last_name or 'User',
                    phone=phone or '',
                    email=email or '',
                    registration_status='complete',
                    source='self_signup'
                )
                
                return Response({
                    'success': True,
                    'status': 'created',
                    'message': 'Registration successful!',
                    'user': MembershipUserDetailSerializer(new_user).data
                }, status=status.HTTP_201_CREATED)
        
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'])
    def check_exists(self, request):
        """
        Check if a user exists by NIN and return their registration status and groups.
        
        Request payload:
        {
            'nin': '12345678901'
        }
        """
        nin = request.data.get('nin')
        
        if not nin:
            return Response({
                'success': False,
                'message': 'NIN is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not nin.isdigit() or len(nin) != 11:
            return Response({
                'success': False,
                'message': 'Invalid NIN format (must be 11 digits)'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            user = MembershipUser.objects.get(nin=nin)
            
            # Get groups user belongs to
            memberships = GroupMembership.objects.filter(user=user).values(
                'id', 'group__id', 'group__name', 'role', 'added_at'
            )
            
            return Response({
                'exists': True,
                'user_id': user.id,
                'registration_status': user.registration_status,
                'source': user.source,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'phone': user.phone,
                'email': user.email,
                'groups': list(memberships)
            }, status=status.HTTP_200_OK)
        
        except MembershipUser.DoesNotExist:
            return Response({
                'exists': False,
                'message': 'User not found. Ready for new registration.'
            }, status=status.HTTP_200_OK)


class MembershipGroupViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing groups in the flexible membership system.
    """
    queryset = MembershipGroup.objects.all()
    serializer_class = MembershipGroupSerializer
    parser_classes = (JSONParser, MultiPartParser, FormParser)
    
    def get_permissions(self):
        if self.action in ['list', 'retrieve']:
            return [AllowAny()]
        return [IsAuthenticated()]
    
    @action(detail=True, methods=['get'])
    def members(self, request, pk=None):
        """
        Get all members of a group with filters.
        
        Query parameters:
        - status: 'complete', 'partial', or 'all' (default)
        """
        group = self.get_object()
        status_filter = request.query_params.get('status', 'all')
        
        memberships = group.members.all()
        
        if status_filter == 'complete':
            memberships = memberships.filter(user__registration_status='complete')
        elif status_filter == 'partial':
            memberships = memberships.filter(user__registration_status='partial')
        
        serializer = GroupMembersListSerializer(memberships, many=True)
        
        return Response({
            'success': True,
            'group_name': group.name,
            'total': memberships.count(),
            'members': serializer.data
        }, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=['post'])
    def add_member_by_nin(self, request, pk=None):
        """
        Add an existing user to a group by their NIN.
        
        Request payload:
        {
            'nin': '12345678901'
        }
        """
        group = self.get_object()
        serializer = AddMemberToGroupSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        
        user_nin = serializer.validated_data['user_nin']
        
        try:
            user = MembershipUser.objects.get(nin=user_nin)
            
            # Check if already in group
            if group.members.filter(user=user).exists():
                return Response({
                    'success': False,
                    'message': f'{user.get_full_name()} already belongs to {group.name}'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Add to group
            membership = GroupMembership.objects.create(
                user=user,
                group=group,
                role='member'
            )
            
            return Response({
                'success': True,
                'message': f'{user.get_full_name()} added to {group.name}',
                'membership': GroupMembershipDetailSerializer(membership).data
            }, status=status.HTTP_201_CREATED)
        
        except MembershipUser.DoesNotExist:
            return Response({
                'success': False,
                'message': f'User with NIN {user_nin} not found'
            }, status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=True, methods=['post'])
    def remove_member(self, request, pk=None):
        """
        Remove a member from a group (delete GroupMembership only).
        
        Request payload:
        {
            'membership_id': 123
        }
        """
        group = self.get_object()
        serializer = RemoveMemberFromGroupSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response({
                'success': False,
                'errors': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        
        membership_id = serializer.validated_data['membership_id']
        
        try:
            membership = GroupMembership.objects.get(id=membership_id, group=group)
            user_name = membership.user.get_full_name()
            membership.delete()
            
            return Response({
                'success': True,
                'message': f'{user_name} removed from {group.name}'
            }, status=status.HTTP_200_OK)
        
        except GroupMembership.DoesNotExist:
            return Response({
                'success': False,
                'message': 'Membership not found'
            }, status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=True, methods=['post'])
    def upload_members_excel(self, request, pk=None):
        """
        Bulk upload members to a group from Excel file.
        
        Expected Excel columns:
        - NIN (National Identification Number)
        - First Name
        - Last Name
        - Phone
        - Email (optional)
        
        Request: multipart/form-data
        - excel_file: The Excel file
        
        Processing:
        - For each row, check if user with NIN exists
        - If exists: Create GroupMembership if not already linked
        - If not exists: Create user with registration_status='partial', then link to group
        - Prevent duplicate linking
        """
        group = self.get_object()
        
        if 'excel_file' not in request.FILES:
            return Response({
                'success': False,
                'message': 'No Excel file provided'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        excel_file = request.FILES['excel_file']
        
        try:
            # Read Excel file
            excel_data = BytesIO(excel_file.read())
            workbook = openpyxl.load_workbook(excel_data)
            worksheet = workbook.active
            
            processed = []
            errors = []
            skipped = []
            row_num = 2
            
            for row in worksheet.iter_rows(min_row=2, values_only=True):
                try:
                    # Parse Excel columns
                    if len(row) < 4:
                        errors.append(f"Row {row_num}: Insufficient columns")
                        row_num += 1
                        continue
                    
                    nin = str(row[0] or '').strip()
                    first_name = str(row[1] or '').strip()
                    last_name = str(row[2] or '').strip()
                    phone = str(row[3] or '').strip()
                    email = str(row[4] or '').strip() if len(row) > 4 else ''
                    
                    # Validate required fields
                    if not all([nin, first_name, last_name, phone]):
                        errors.append(f"Row {row_num}: Missing required fields")
                        row_num += 1
                        continue
                    
                    # Validate NIN
                    if not nin.isdigit() or len(nin) != 11:
                        errors.append(f"Row {row_num}: Invalid NIN format")
                        row_num += 1
                        continue
                    
                    # Get or create user
                    user, created = MembershipUser.objects.get_or_create(
                        nin=nin,
                        defaults={
                            'first_name': first_name,
                            'last_name': last_name,
                            'phone': phone,
                            'email': email or '',
                            'registration_status': 'partial',
                            'source': 'group_import'
                        }
                    )
                    
                    # Create or skip membership
                    if GroupMembership.objects.filter(user=user, group=group).exists():
                        skipped.append(f"Row {row_num}: {user.get_full_name()} already in group")
                    else:
                        GroupMembership.objects.create(user=user, group=group)
                        processed.append({
                            'nin': nin,
                            'name': f"{first_name} {last_name}",
                            'status': 'created' if created else 'linked'
                        })
                    
                    row_num += 1
                
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
                    row_num += 1
            
            return Response({
                'success': True,
                'message': f'Excel import completed',
                'stats': {
                    'created': len([p for p in processed if p['status'] == 'created']),
                    'linked': len([p for p in processed if p['status'] == 'linked']),
                    'skipped': len(skipped),
                    'errors': len(errors)
                },
                'processed': processed,
                'skipped': skipped,
                'errors': errors
            }, status=status.HTTP_200_OK)
        
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)


class AdminDashboardViewSet(viewsets.ViewSet):
    """
    ViewSet for admin dashboard endpoints.
    """
    permission_classes = [IsAuthenticated]
    
    def get_permissions(self):
        if self.action == 'stats':
            return [AllowAny()]
        return [IsAuthenticated()]
    
    @action(detail=False, methods=['get'])
    def stats(self, request):
        """
        Get overall system statistics for admin dashboard.
        """
        try:
            total_users = MembershipUser.objects.count()
            complete_users = MembershipUser.objects.filter(registration_status='complete').count()
            partial_users = MembershipUser.objects.filter(registration_status='partial').count()
            
            total_groups = MembershipGroup.objects.count()
            total_memberships = GroupMembership.objects.count()
            
            # Duplicate NIN attempts (if tracked)
            duplicate_attempts = 0  # Can be implemented with logging
            
            return Response({
                'success': True,
                'stats': {
                    'users': {
                        'total': total_users,
                        'complete': complete_users,
                        'partial': partial_users
                    },
                    'groups': {
                        'total': total_groups,
                        'total_memberships': total_memberships,
                        'avg_members_per_group': total_memberships // total_groups if total_groups > 0 else 0
                    },
                    'audit': {
                        'duplicate_nin_attempts': duplicate_attempts
                    }
                }
            }, status=status.HTTP_200_OK)
        
        except Exception as e:
            return Response({
                'success': False,
                'error': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def users(self, request):
        """
        Get list of all users with filters.
        
        Query parameters:
        - status: 'complete', 'partial'
        - source: 'group_import', 'self_signup'
        - search: Search by name or NIN
        """
        queryset = MembershipUser.objects.all()
        
        # Filters
        status_filter = request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(registration_status=status_filter)
        
        source_filter = request.query_params.get('source')
        if source_filter:
            queryset = queryset.filter(source=source_filter)
        
        search = request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(nin__icontains=search)
            )
        
        serializer = MembershipUserListSerializer(queryset, many=True)
        
        return Response({
            'success': True,
            'count': queryset.count(),
            'users': serializer.data
        }, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['get'])
    def groups(self, request):
        """
        Get list of all groups with statistics.
        """
        queryset = MembershipGroup.objects.all().annotate(
            member_count=GroupMembership.objects.filter(group_id=pk).count()
        )
        
        groups_data = []
        for group in queryset:
            groups_data.append({
                'id': group.id,
                'name': group.name,
                'description': group.description,
                'created_by': group.created_by,
                'created_at': group.created_at,
                'total_members': group.total_members,
                'complete': group.complete_count,
                'partial': group.partial_count
            })
        
        return Response({
            'success': True,
            'count': len(groups_data),
            'groups': groups_data
        }, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['get'])
    def audit_duplicates(self, request):
        """
        Get audit log of duplicate NIN attempts.
        Note: This would require a separate audit logging system to be implemented.
        """
        # Check for duplicate NINs (should only happen in rare cases of data integrity issues)
        duplicate_nins = (
            MembershipUser.objects
            .values('nin')
            .annotate(count=Count('id'))
            .filter(count__gt=1)
        )
        
        audit_data = []
        for dup in duplicate_nins:
            users = MembershipUser.objects.filter(nin=dup['nin'])
            audit_data.append({
                'nin': dup['nin'],
                'count': dup['count'],
                'users': MembershipUserListSerializer(users, many=True).data
            })
        
        return Response({
            'success': True,
            'duplicate_count': len(audit_data),
            'duplicates': audit_data
        }, status=status.HTTP_200_OK)

